"""Excel spreadsheet import into Flow.

Usage (via flow_cli.py):
    flow import <file.xlsx> [--config config/import_config.yaml] [--preview] [--dry-run]
"""

import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None


_REPO_ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
OUTPUTS_DIR = os.path.join(_REPO_ROOT, "outputs")


def _require_openpyxl():
    if openpyxl is None:
        print("ERROR: openpyxl is required for import. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def _load_config(config_path):
    """Load import column mapping config. Returns dict."""
    if config_path and os.path.exists(config_path):
        try:
            import yaml
            with open(config_path) as f:
                return yaml.safe_load(f)
        except ImportError:
            print("WARNING: PyYAML not installed, using default column mapping.", file=sys.stderr)
    # Default mapping: first row is headers, map by common column names
    return {
        "column_map": {
            "name": ["name", "id", "req id", "requirement id"],
            "description": ["description", "statement", "text", "requirement"],
            "owner": ["owner", "responsible", "assignee"],
            "type": ["type", "req type", "requirement type"],
            "rationale": ["rationale", "justification"],
            "notes": ["notes", "comments"],
        },
        "sheet_to_system": {},  # sheet name -> system id override
        "skip_empty_rows": True,
    }


def _map_columns(headers, column_map):
    """Return a dict mapping field_name -> column_index (0-based)."""
    mapping = {}
    lower_headers = [str(h).lower().strip() if h else "" for h in headers]
    for field, candidates in column_map.items():
        for candidate in candidates:
            if candidate.lower() in lower_headers:
                mapping[field] = lower_headers.index(candidate.lower())
                break
    return mapping


def _import_id(sheet_name, row_index):
    return f"{sheet_name}::row{row_index}"


def parse_workbook(xlsx_path, config):
    """Parse the workbook and return a list of parsed requirement dicts."""
    _require_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    column_map = config.get("column_map", {})
    skip_empty = config.get("skip_empty_rows", True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = rows[0]
        col_map = _map_columns(headers, column_map)

        if "name" not in col_map and "description" not in col_map:
            print(f"  WARNING: Sheet '{sheet_name}' - could not find name or description columns. Skipping.", file=sys.stderr)
            continue

        for row_idx, row in enumerate(rows[1:], start=2):
            if skip_empty and all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            req = {
                "import_id": _import_id(sheet_name, row_idx),
                "sheet": sheet_name,
                "row": row_idx,
            }
            for field, col_idx in col_map.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    req[field] = str(val).strip() if val is not None else ""
                else:
                    req[field] = ""
            # Ensure name falls back to import_id if blank
            if not req.get("name"):
                req["name"] = req["import_id"]
            results.append(req)

    return results


def check_existing_import_ids(client_module, org, project):
    """Fetch all requirements and return a set of existing import_ids from notes/custom fields."""
    reqs = client_module.api_request_all_pages(f"/org/{org}/project/{project}/requirements/paged")
    existing = set()
    for r in reqs:
        # import_id is stored in the notes field with prefix "import_id:"
        notes = r.get("notes", "") or ""
        for line in notes.splitlines():
            if line.startswith("import_id:"):
                existing.add(line.split(":", 1)[1].strip())
    return existing


def run_import(args):
    """Main import entrypoint."""
    from flow import client, backup, quality as qual_module

    client.load_dotenv()
    org, project = client.get_org_and_project(args)

    config_path = getattr(args, "config", None) or os.path.join(_REPO_ROOT, "config", "import_config.yaml")
    config = _load_config(config_path)

    xlsx_path = args.file
    if not os.path.exists(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing: {xlsx_path}")
    rows = parse_workbook(xlsx_path, config)
    print(f"  Found {len(rows)} rows across all sheets.")

    preview = getattr(args, "preview", False)
    dry_run = getattr(args, "dry_run", False)

    if preview or dry_run:
        print("\n--- PREVIEW ---")
        print(f"{'Sheet':<20} {'Row':>4}  {'Name':<40}  {'Description':<60}")
        print("-" * 130)
        for r in rows[:50]:
            print(f"{r['sheet']:<20} {r['row']:>4}  {r.get('name','')[:40]:<40}  {r.get('description','')[:60]:<60}")
        if len(rows) > 50:
            print(f"  ... and {len(rows) - 50} more rows")

    if dry_run:
        print("\n--- DRY RUN: Quality check ---")
        issues = []
        for r in rows:
            desc = r.get("description", "")
            if not desc or "shall" not in desc.lower():
                issues.append((r["import_id"], "INC001", "Statement does not contain 'shall'"))
        if issues:
            for import_id, rule_id, msg in issues:
                print(f"  [WARNING] {import_id}: {rule_id} - {msg}")
        else:
            print("  No quality issues found.")
        return

    if preview:
        return

    # Full import
    print("\nChecking for already-imported rows...")
    existing_ids = check_existing_import_ids(client, org, project)
    to_import = [r for r in rows if r["import_id"] not in existing_ids]
    skipped = len(rows) - len(to_import)
    print(f"  {skipped} rows already imported (skipping), {len(to_import)} new rows to import.")

    if not to_import:
        print("Nothing to import.")
        return

    # Backup current state
    print("Creating backup...")
    current_reqs = client.api_request_all_pages(f"/org/{org}/project/{project}/requirements/paged")
    backup_id = backup.create_backup("import", current_reqs, manifest={"file": xlsx_path, "rows_to_import": len(to_import)})

    # Create requirements
    path = f"/org/{org}/project/{project}/requirements"
    created = []
    failed = []
    for r in to_import:
        notes_tag = f"import_id:{r['import_id']}"
        existing_notes = r.get("notes", "")
        full_notes = f"{notes_tag}\n{existing_notes}".strip()
        body = [{"name": r.get("name", ""), "description": r.get("description", ""), "notes": full_notes}]
        try:
            result = client.api_request("POST", path, body)
            created.append({"import_id": r["import_id"], "result": result})
            print(f"  Created: {r['import_id']}")
        except SystemExit:
            failed.append(r["import_id"])

    # Write report
    os.makedirs(OUTPUTS_DIR, exist_ok=True)
    report_path = os.path.join(OUTPUTS_DIR, f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    with open(report_path, "w") as f:
        json.dump({"backup_id": backup_id, "created": len(created), "skipped": skipped, "failed": failed, "rows": created}, f, indent=2)

    print(f"\nImport complete: {len(created)} created, {skipped} skipped, {len(failed)} failed.")
    print(f"Report: {report_path}")
    print(f"Backup ID: {backup_id}  (use 'flow restore {backup_id}' to undo)")
